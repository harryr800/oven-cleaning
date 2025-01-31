import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_brrrr_excel(data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BRRRR Calculator"

    # Formatting Helpers
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center")

    # Helper to create boxed sections
    def create_box(start_row, start_col, title, fields, values, is_formula=False):
        # Section Title
        sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 2)
        cell = sheet.cell(row=start_row, column=start_col, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = center_align

        # Add Fields and Values
        for i, (field, value) in enumerate(zip(fields, values), start=1):
            field_cell = sheet.cell(row=start_row + i, column=start_col, value=field)
            field_cell.font = bold_font
            field_cell.alignment = center_align
            field_cell.border = border

            value_cell = sheet.cell(row=start_row + i, column=start_col + 1, value=value if not is_formula else f"={value}")
            value_cell.fill = light_yellow_fill
            value_cell.alignment = center_align
            value_cell.border = border

    # BUY THE PROPERTY Section
    buy_fields = [
        "Purchased Price", "Down Payment %", "Mortgage Amount", "Closing Cost", "Interest Rate",
        "Amortization", "Monthly Payment", "Months of Carrying Cost", "Total Cost"
    ]
    buy_values = [
        data['Purchase Price'], f"{data['Down Payment %']}%", f"B2*(1-B3/100)", f"B2*2%", f"{data['Interest Rate']}%",
        30, f"B4*(B5/12)", data['Carrying Cost Months'], "B4+B6+B7+B8"
    ]
    create_box(1, 1, "BUY THE PROPERTY", buy_fields, buy_values, is_formula=True)

    # REFINANCE POST RENO SECTION WITH PMT FORMULA
    refinance_fields = [
        "New Property Value", "New Mortgage Amount LTV", "Old Mortgage Balance", "Penalty", "Lawyer Fees",
        "Interest Rate New Mortgage", "Amortization New Mortgage", "Monthly Payment New Mortgage"
    ]
    refinance_values = [
        data['Post Reno Value'], f"B14*B15/100", "B4", 6965, 2000, f"{data['Interest Rate']}%", 30, "PMT(B23/12, B24*12, -B22)"
    ]
    create_box(20, 5, "REFINANCE POST RENO", refinance_fields, refinance_values, is_formula=True)

    # FUTURE PROPERTY VALUE TABLE WITH AMORTIZATION
    sheet.cell(row=30, column=1, value="FUTURE PROPERTY VALUE WITH AMORTIZATION").font = Font(bold=True, color="FFFFFF")
    sheet.cell(row=30, column=1).fill = header_fill

    amort_fields = ["Year", "Payment", "Interest", "Principal", "Balance"]
    for col_idx, field in enumerate(amort_fields, start=1):
        sheet.cell(row=31, column=col_idx, value=field).font = bold_font

    # Corrected Amortization Table with References
    balance = "B22"
    for year in range(1, 6):
        row = 31 + year
        sheet.cell(row=row, column=1, value=year)  # Year
        sheet.cell(row=row, column=2, value="B24")  # Payment (constant)
        sheet.cell(row=row, column=3, value=f"{balance}*B23/12")  # Interest
        sheet.cell(row=row, column=4, value=f"B24-B{row}C")  # Principal
        balance = f"{balance}-B{row}D"  # Updated Balance
        sheet.cell(row=row, column=5, value=f"{balance}")

    # Auto-adjust column widths
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20

    # Save Workbook
    workbook.save("10-BRRRR_Property_Calculator.xlsx")
    print("10-BRRRR Property Calculator with Corrected Formulas Created!")

if __name__ == "__main__":
    # Input data
    data = {
        'Purchase Price': 1050000,
        'Down Payment %': 20,
        'Interest Rate': 3.4,
        'Carrying Cost Months': 8,
        'Renovation Cost': 128000,
        'Post Reno Value': 1350000
    }
    create_brrrr_excel(data)
