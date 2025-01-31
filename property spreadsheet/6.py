import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_brrrr_excel(data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BRRRR Calculator"

    # Formatting Helpers
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    center_align = Alignment(horizontal="center")
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Function to write input table with Excel formulas
    def write_input_table(start_row, start_col, section_name, fields, values):
        sheet.cell(row=start_row, column=start_col, value=section_name).font = bold_font
        sheet.cell(row=start_row, column=start_col).alignment = center_align
        sheet.cell(row=start_row, column=start_col).fill = header_fill

        for i, field in enumerate(fields, start=1):
            sheet.cell(row=start_row + i, column=start_col, value=field).font = bold_font
            sheet.cell(row=start_row + i, column=start_col).alignment = center_align
            sheet.cell(row=start_row + i, column=start_col + 1, value=values[i - 1])
            sheet.cell(row=start_row + i, column=start_col + 1).border = border

    # Function to write output table with Excel formulas
    def write_output_table(start_row, start_col, section_name, fields, formulas):
        sheet.cell(row=start_row, column=start_col, value=section_name).font = bold_font
        sheet.cell(row=start_row, column=start_col).alignment = center_align
        sheet.cell(row=start_row, column=start_col).fill = header_fill

        for i, field in enumerate(fields, start=1):
            sheet.cell(row=start_row + i, column=start_col, value=field).font = bold_font
            sheet.cell(row=start_row + i, column=start_col).alignment = center_align
            sheet.cell(row=start_row + i, column=start_col + 1, value=f"={formulas[i - 1]}")
            sheet.cell(row=start_row + i, column=start_col + 1).border = border

    # Input Table Fields and Values
    input_fields = [
        "Property Address",
        "Property Type",
        "Square Footage",
        "Number of Units",
        "Purchase Price",
        "Down Payment %",
        "Interest Rate",
        "Amortization Years",
        "Closing Cost",
        "Renovation Cost",
        "Carrying Cost Months",
        "Post Reno Value",
        "Refinance LTV %",
        "Appreciation Rate %"
    ]
    input_values = [
        data.get('Property Address', ""),
        data.get('Property Type', ""),
        data.get('Square Footage', ""),
        data.get('Number of Units', ""),
        data['Purchase Price'],
        data['Down Payment %'],
        data['Interest Rate'],
        data['Amortization Years'],
        data['Closing Cost'],
        data['Renovation Cost'],
        data['Carrying Cost Months'],
        data['Post Reno Value'],
        data['Refinance LTV'],
        data['Appreciation Rate']
    ]
    write_input_table(1, 1, "INPUT FIELDS", input_fields, input_values)

    # Output Table Fields and Formulas
    output_fields = [
        "Down Payment",
        "Mortgage Amount",
        "Monthly Payment",
        "Carrying Cost",
        "Total Cost",
        "Post Reno Income",
        "Total Expenses",
        "Monthly Cashflow",
        "Cash Pulled Out",
        "ROI (%)"
    ]
    output_formulas = [
        "B5*B6/100",  # Down Payment = Purchase Price * Down Payment %
        "B5-B17",     # Mortgage Amount = Purchase Price - Down Payment
        "(B18*B7/12)", # Monthly Payment = Mortgage Amount * Interest Rate / 12
        "B19*B11",     # Carrying Cost = Monthly Payment * Carrying Cost Months
        "B6+B9+B10+B20", # Total Cost = Down Payment + Closing Cost + Renovation Cost + Carrying Cost
        "SUM(B21:B23)", # Post Reno Income = SUM Rent Income
        "SUM(B26:B28)", # Total Expenses = SUM Operating Expenses
        "B25-B30",     # Monthly Cashflow = Post Reno Income - Total Expenses
        "(B13*B12/100)-B18", # Cash Pulled Out = Refinance Value - Mortgage Amount
        "(B31*12)/B20*100"  # ROI = Annual Cashflow / Total Cost * 100
    ]
    write_output_table(1, 5, "OUTPUT FIELDS", output_fields, output_formulas)

    # Auto-adjust column widths
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20

    # Save Workbook
    workbook.save("6-BRRRR_Property_Calculator.xlsx")
    print("BRRRR Property Calculator with Excel Formulas has been created!")

if __name__ == "__main__":
    # Input data
    data = {
        'Property Address': "123 Main St",
        'Property Type': "Residential",
        'Square Footage': 1500,
        'Number of Units': 1,
        'Purchase Price': 1050000,
        'Down Payment %': 20,
        'Interest Rate': 3.4,
        'Carrying Cost Months': 8,
        'Closing Cost': 21000,
        'Renovation Cost': 128000,
        'Post Reno Value': 1350000,
        'Refinance LTV': 80,
        'Amortization Years': 5,
        'Appreciation Rate': 8
    }
    create_brrrr_excel(data)
