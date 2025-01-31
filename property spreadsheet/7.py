import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_brrrr_excel(data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BRRRR Calculator"

    # Formatting Helpers
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    center_align = Alignment(horizontal="center")
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Function to write input/output tables
    def write_table(start_row, start_col, section_name, fields, values, formulas=False):
        sheet.cell(row=start_row, column=start_col, value=section_name).font = bold_font
        sheet.cell(row=start_row, column=start_col).alignment = center_align
        sheet.cell(row=start_row, column=start_col).fill = header_fill

        for i, field in enumerate(fields, start=1):
            sheet.cell(row=start_row + i, column=start_col, value=field).font = bold_font
            sheet.cell(row=start_row + i, column=start_col).alignment = center_align
            value = values[i - 1] if not formulas else f"={values[i - 1]}"
            sheet.cell(row=start_row + i, column=start_col + 1, value=value).border = border

    # BUY THE PROPERTY Section
    buy_fields = [
        "Purchased Price", "Down Payment %", "Mortgage Amount", "Closing Cost", "Interest Rate",
        "Amortization", "Monthly Payment", "Months of Carrying Cost"
    ]
    buy_values = [
        data['Purchase Price'], f"{data['Down Payment %']}%", f"B2*(1-B3/100)", f"B2*2%", f"{data['Interest Rate']}%",
        30, f"B4*(B5/12)", data['Carrying Cost Months']
    ]
    write_table(1, 1, "BUY THE PROPERTY", buy_fields, buy_values, formulas=True)

    # POST RENO OPERATING EXPENSES
    expense_fields = [
        "Hydro", "Water", "Gas", "Insurance", "Property Tax", "Vacancy %", "Maintenance %", "Property Management %"
    ]
    expense_values = [50, 50, 100, 150, 291, "5%", "5%", "10%"]
    write_table(1, 5, "POST RENO OPERATING EXPENSES", expense_fields, expense_values)

    # REHAB THE PROPERTY
    rehab_fields = ["Renovation Cost", "Appliances", "Construction Insurance"]
    rehab_values = [data['Renovation Cost'], 8000, 0]
    write_table(10, 1, "REHAB THE PROPERTY", rehab_fields, rehab_values)

    # RENT MONTHLY INCOME POST RENO
    rent_fields = ["Rent", "Parking", "Laundry"]
    rent_values = [3200, 350, 0]
    write_table(10, 5, "RENT MONTHLY INCOME POST RENO", rent_fields, rent_values)

    # SUMMARY
    summary_fields = [
        "Upfront Investment", "Cash Pulled Out After Refinancing", "Money Left In Property", "Percent Down (LEVERAGE)",
        "Mortgage Paydown After 5 Years", "Equity in Property After 5 Years"
    ]
    summary_values = [
        "B2*B3/100+B6+B7", "(B10*B11/100)-B4", "B2-B12", "B12/B2", 104299, 1007892
    ]
    write_table(15, 1, "SUMMARY", summary_fields, summary_values, formulas=True)

    # FUTURE PROPERTY VALUE TABLE
    amort_start = 20
    sheet.cell(row=amort_start, column=1, value="Year").font = bold_font
    sheet.cell(row=amort_start, column=2, value="Future Property Value").font = bold_font
    for year in range(6):
        sheet.cell(row=amort_start + year + 1, column=1, value=year)
        sheet.cell(row=amort_start + year + 1, column=2, value=f"=B10*(1+8%)^{year}").alignment = center_align

    # Auto-adjust column widths
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20

    # Save Workbook
    workbook.save("7-BRRRR_Property_Calculator.xlsx")
    print("BRRRR Property Calculator with Formulas and Layout has been created!")

if __name__ == "__main__":
    # Input data
    data = {
        'Purchase Price': 1050000,
        'Down Payment %': 20,
        'Interest Rate': 3.4,
        'Carrying Cost Months': 8,
        'Renovation Cost': 128000
    }
    create_brrrr_excel(data)
