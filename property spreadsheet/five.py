import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def calculate_brrrr_metrics(data):
    # Input fields
    purchase_price = data['Purchase Price']
    down_payment_percentage = data['Down Payment %'] / 100
    mortgage_amount = purchase_price * (1 - down_payment_percentage)
    interest_rate = data['Interest Rate'] / 100
    monthly_payment = mortgage_amount * (interest_rate / 12)
    carrying_cost_months = data['Carrying Cost Months']
    carrying_cost = monthly_payment * carrying_cost_months
    closing_cost = data['Closing Cost']
    renovation_cost = data['Renovation Cost']
    total_cost = purchase_price * down_payment_percentage + closing_cost + carrying_cost + renovation_cost

    # Monthly rent
    post_reno_income = sum(data['Rent Income'].values())
    operating_expenses = data['Operating Expenses']
    total_expenses = sum(operating_expenses.values())
    monthly_cashflow = post_reno_income - total_expenses

    # Refinancing calculations
    post_reno_value = data['Post Reno Value']
    refinance_amount = post_reno_value * data['Refinance LTV'] / 100
    cash_out = refinance_amount - mortgage_amount
    roi = (monthly_cashflow * 12) / total_cost * 100

    # Future value calculations
    amortization_years = data['Amortization Years']
    appreciation_rate = data['Appreciation Rate'] / 100
    future_value = [post_reno_value * (1 + appreciation_rate) ** year for year in range(amortization_years + 1)]
    payment = monthly_payment * 12
    principal_balance = mortgage_amount
    amortization_schedule = []

    for year in range(amortization_years):
        interest = principal_balance * interest_rate
        principal = payment - interest
        principal_balance -= principal
        amortization_schedule.append((payment, interest, principal, principal_balance))

    return {
        'Purchase Price': purchase_price,
        'Down Payment': purchase_price * down_payment_percentage,
        'Mortgage Amount': mortgage_amount,
        'Monthly Payment': monthly_payment,
        'Carrying Cost': carrying_cost,
        'Closing Cost': closing_cost,
        'Renovation Cost': renovation_cost,
        'Total Cost': total_cost,
        'Post Reno Income': post_reno_income,
        'Total Expenses': total_expenses,
        'Monthly Cashflow': monthly_cashflow,
        'Cash Pulled Out': cash_out,
        'ROI': roi,
        'Future Property Values': future_value,
        'Amortization Schedule': amortization_schedule
    }

def create_brrrr_excel(data):
    metrics = calculate_brrrr_metrics(data)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BRRRR Calculator"

    # Formatting Helpers
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    center_align = Alignment(horizontal="center")
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Function to write data neatly
    def write_table(start_row, start_col, section_name, fields, values=None):
        sheet.cell(row=start_row, column=start_col, value=section_name).font = bold_font
        sheet.cell(row=start_row, column=start_col).alignment = center_align
        sheet.cell(row=start_row, column=start_col).fill = header_fill
        for i, field in enumerate(fields, start=1):
            sheet.cell(row=start_row + i, column=start_col, value=field).font = bold_font
            sheet.cell(row=start_row + i, column=start_col).alignment = center_align
            value = values[i - 1] if values else ""
            sheet.cell(row=start_row + i, column=start_col + 1, value=value).border = border

    # Input Table Fields and Values
    input_fields = [
        "Property Address",
        "Property Type",
        "Square Footage",
        "Number of Units",
        "Purchase Price",
        "Down Payment %",
        "Interest Rate",
        "Amortization Years",
        "Closing Cost",
        "Renovation Cost",
        "Carrying Cost Months",
        "Post Reno Value",
        "Refinance LTV %",
        "Appreciation Rate %"
    ]
    input_values = [
        data.get('Property Address', ""),
        data.get('Property Type', ""),
        data.get('Square Footage', ""),
        data.get('Number of Units', ""),
        data['Purchase Price'],
        data['Down Payment %'],
        data['Interest Rate'],
        data['Amortization Years'],
        data['Closing Cost'],
        data['Renovation Cost'],
        data['Carrying Cost Months'],
        data['Post Reno Value'],
        data['Refinance LTV'],
        data['Appreciation Rate']
    ]
    write_table(1, 1, "INPUT FIELDS", input_fields, input_values)

    # Output Table Fields
    output_fields = [
        "Purchase Price",
        "Down Payment",
        "Mortgage Amount",
        "Monthly Payment",
        "Carrying Cost",
        "Closing Cost",
        "Renovation Cost",
        "Total Cost",
        "Post Reno Income",
        "Total Expenses",
        "Monthly Cashflow",
        "Cash Pulled Out",
        "ROI (%)"
    ]
    output_values = [
        metrics['Purchase Price'],
        metrics['Down Payment'],
        metrics['Mortgage Amount'],
        metrics['Monthly Payment'],
        metrics['Carrying Cost'],
        metrics['Closing Cost'],
        metrics['Renovation Cost'],
        metrics['Total Cost'],
        metrics['Post Reno Income'],
        metrics['Total Expenses'],
        metrics['Monthly Cashflow'],
        metrics['Cash Pulled Out'],
        metrics['ROI']
    ]
    write_table(1, 5, "OUTPUT FIELDS", output_fields, output_values)

    # Write Data Sections
    write_table(20, 1, "OPERATING EXPENSES", list(data['Operating Expenses'].keys()), list(data['Operating Expenses'].values()))
    write_table(20, 5, "RENTAL INCOME SOURCES", list(data['Rent Income'].keys()), list(data['Rent Income'].values()))

    # Auto-adjust column widths
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20

    # Save Workbook
    workbook.save("BRRRR_Property_Calculator.xlsx")
    print("BRRRR Property Calculator has been created with Input and Output Fields populated!")

if __name__ == "__main__":
    # Input data
    data = {
        'Property Address': "123 Main St",
        'Property Type': "Residential",
        'Square Footage': 1500,
        'Number of Units': 1,
        'Purchase Price': 1050000,
        'Down Payment %': 20,
        'Interest Rate': 3.4,
        'Carrying Cost Months': 8,
        'Closing Cost': 21000,
        'Renovation Cost': 128000,
        'Post Reno Value': 1350000,
        'Refinance LTV': 80,
        'Amortization Years': 5,
        'Appreciation Rate': 8,
        'Rent Income': {"Rent 1": 3200, "Parking": 350, "Laundry": 0},
        'Operating Expenses': {
            "Hydro": 50, "Water": 50, "Gas": 100, "Insurance": 150, "Property Tax": 291,
            "Vacancy": 388, "Maintenance": 388, "Property Management": 775
        }
    }
    create_brrrr_excel(data)
