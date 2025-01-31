import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_brrrr_excel(data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BRRRR Calculator"

    # Formatting Helpers
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center")

    # Helper to create boxed sections
    def create_box(start_row, start_col, title, fields, values, is_formula=False):
        # Section Title
        sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 2)
        cell = sheet.cell(row=start_row, column=start_col, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = center_align

        # Add Fields and Values
        for i, (field, value) in enumerate(zip(fields, values), start=1):
            field_cell = sheet.cell(row=start_row + i, column=start_col, value=field)
            field_cell.font = bold_font
            field_cell.alignment = center_align
            field_cell.border = border

            value_cell = sheet.cell(row=start_row + i, column=start_col + 1, value=value if not is_formula else f"={value}")
            value_cell.fill = light_yellow_fill
            value_cell.alignment = center_align
            value_cell.border = border

    # BUY THE PROPERTY Section
    buy_fields = [
        "Purchased Price", "Down Payment %", "Mortgage Amount", "Closing Cost", "Interest Rate",
        "Amortization", "Monthly Payment", "Months of Carrying Cost", "Total Cost"
    ]
    buy_values = [
        data['Purchase Price'], f"{data['Down Payment %']}%", f"B2*(1-B3/100)", f"B2*2%", f"{data['Interest Rate']}%",
        30, f"B4*(B5/12)", data['Carrying Cost Months'], "B4+B6+B7+B8"
    ]
    create_box(1, 1, "BUY THE PROPERTY", buy_fields, buy_values, is_formula=True)

    # POST RENO OPERATING EXPENSES
    expense_fields = [
        "Hydro", "Water", "Gas", "Insurance", "Property Tax", "Vacancy", "Maintenance", "Property Management"
    ]
    expense_values = [50, 50, 100, 150, 291, "5%", "5%", "10%"]
    create_box(1, 5, "POST RENO OPERATING EXPENSES", expense_fields, expense_values)

    # REHAB THE PROPERTY
    rehab_fields = ["Renovation Cost", "Appliances", "Construction Insurance", "Total Cost"]
    rehab_values = [data['Renovation Cost'], 8000, 0, "B11+B12+B13"]
    create_box(12, 1, "REHAB THE PROPERTY", rehab_fields, rehab_values, is_formula=True)

    # RENT MONTHLY INCOME POST RENO
    rent_fields = ["Rent", "Parking", "Laundry", "Total Income"]
    rent_values = [3200, 350, 0, "SUM(B18:B20)"]
    create_box(12, 5, "RENT MONTHLY INCOME POST RENO", rent_fields, rent_values, is_formula=True)

    # SUMMARY SECTION
    summary_fields = [
        "Upfront Investment", "Cash Pulled Out After Refinancing", "Money Left In Property",
        "Percent Down (LEVERAGE)", "Mortgage Paydown After 5 Years", "Equity in Property After 5 Years"
    ]
    summary_values = [
        "B4+B11+B12+B13", "(B14*B15/100)-B6", "B4-B17", "B17/B4", 104299, 1007892
    ]
    create_box(20, 1, "SUMMARY", summary_fields, summary_values, is_formula=True)

    # FUTURE PROPERTY VALUE TABLE
    sheet.cell(row=20, column=5, value="FUTURE PROPERTY VALUE").font = Font(bold=True, color="FFFFFF")
    sheet.cell(row=20, column=5).fill = header_fill
    sheet.cell(row=21, column=5, value="Year").font = bold_font
    sheet.cell(row=21, column=6, value="Value").font = bold_font

    for year in range(6):
        sheet.cell(row=22 + year, column=5, value=year)
        sheet.cell(row=22 + year, column=6, value=f"=B14*(1+8%)^{year}").fill = light_yellow_fill
        sheet.cell(row=22 + year, column=6).alignment = center_align

    # Auto-adjust column widths
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20

    # Save Workbook
    workbook.save("8-BRRRR_Property_Calculator.xlsx")
    print("8-BRRRR Property Calculator with Full Layout, Colors, and Formulas Created!")

if __name__ == "__main__":
    # Input data
    data = {
        'Purchase Price': 1050000,
        'Down Payment %': 20,
        'Interest Rate': 3.4,
        'Carrying Cost Months': 8,
        'Renovation Cost': 128000
    }
    create_brrrr_excel(data)
