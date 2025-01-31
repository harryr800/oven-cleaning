import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def calculate_brrrr_metrics(data):
    # Input fields
    purchase_price = data['Purchase Price']
    down_payment_percentage = data['Down Payment %'] / 100
    mortgage_amount = purchase_price * (1 - down_payment_percentage)
    interest_rate = data['Interest Rate'] / 100
    monthly_payment = mortgage_amount * (interest_rate / 12)
    carrying_cost_months = data['Carrying Cost Months']
    carrying_cost = monthly_payment * carrying_cost_months
    closing_cost = data['Closing Cost']
    renovation_cost = data['Renovation Cost']
    total_cost = down_payment_percentage * purchase_price + closing_cost + carrying_cost + renovation_cost

    # Monthly rent
    post_reno_income = sum(data['Rent Income'].values())
    operating_expenses = data['Operating Expenses']
    total_expenses = sum(operating_expenses.values())
    monthly_cashflow = post_reno_income - total_expenses

    # Refinancing calculations
    post_reno_value = data['Post Reno Value']
    refinance_amount = post_reno_value * data['Refinance LTV'] / 100
    cash_out = refinance_amount - mortgage_amount
    roi = (monthly_cashflow * 12) / total_cost * 100

    # Future value calculations
    amortization_years = data['Amortization Years']
    appreciation_rate = data['Appreciation Rate'] / 100
    future_value = [post_reno_value * (1 + appreciation_rate) ** year for year in range(amortization_years + 1)]
    payment = monthly_payment * 12
    principal_balance = mortgage_amount
    amortization_schedule = []

    for year in range(amortization_years):
        interest = principal_balance * interest_rate
        principal = payment - interest
        principal_balance -= principal
        amortization_schedule.append((payment, interest, principal, principal_balance))

    return {
        'Purchase Price': purchase_price,
        'Down Payment': purchase_price * down_payment_percentage,
        'Mortgage Amount': mortgage_amount,
        'Monthly Payment': monthly_payment,
        'Carrying Cost': carrying_cost,
        'Closing Cost': closing_cost,
        'Renovation Cost': renovation_cost,
        'Total Cost': total_cost,
        'Post Reno Income': post_reno_income,
        'Total Expenses': total_expenses,
        'Monthly Cashflow': monthly_cashflow,
        'Cash Pulled Out': cash_out,
        'ROI': roi,
        'Future Property Values': future_value,
        'Amortization Schedule': amortization_schedule
    }

def create_brrrr_excel(data):
    metrics = calculate_brrrr_metrics(data)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BRRRR Calculator"

    # Formatting Helpers
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    center_align = Alignment(horizontal="center")

    # Function to write data neatly
    def write_section(start_row, start_col, section_name, entries):
        sheet.cell(row=start_row, column=start_col, value=section_name).font = bold_font
        sheet.cell(row=start_row, column=start_col).alignment = center_align
        for i, (key, value) in enumerate(entries.items(), start=1):
            sheet.cell(row=start_row + i, column=start_col, value=key)
            sheet.cell(row=start_row + i, column=start_col + 1, value=value)
            sheet.cell(row=start_row + i, column=start_col).font = bold_font

    # Write Data Sections
    write_section(1, 1, "BUY THE PROPERTY", {
        "Purchased Price": metrics['Purchase Price'],
        "Down Payment": metrics['Down Payment'],
        "Mortgage Amount": metrics['Mortgage Amount'],
        "Monthly Payment": metrics['Monthly Payment'],
        "Carrying Cost": metrics['Carrying Cost'],
        "Closing Cost": metrics['Closing Cost'],
        "Total Cost": metrics['Total Cost'],
    })

    write_section(10, 1, "POST RENO OPERATING EXPENSES", data['Operating Expenses'])
    write_section(10, 5, "RENT OUT THE PROPERTY", {
        "Post Reno Income": metrics['Post Reno Income'],
        "Total Expenses": metrics['Total Expenses'],
        "Monthly Cashflow": metrics['Monthly Cashflow'],
    })

    write_section(20, 1, "SUMMARY", {
        "Cash Pulled Out": metrics['Cash Pulled Out'],
        "ROI": f"{metrics['ROI']}%",
    })

    # Write Amortization Table
    amort_start = 25
    sheet.cell(row=amort_start, column=1, value="Future Property Value with Appreciation").font = bold_font
    sheet.cell(row=amort_start, column=2, value="Payment").font = bold_font
    sheet.cell(row=amort_start, column=3, value="Interest").font = bold_font
    sheet.cell(row=amort_start, column=4, value="Principal").font = bold_font
    sheet.cell(row=amort_start, column=5, value="Balance").font = bold_font

    for year, (payment, interest, principal, balance) in enumerate(metrics['Amortization Schedule'], start=1):
        sheet.cell(row=amort_start + year, column=1, value=year)
        sheet.cell(row=amort_start + year, column=2, value=payment)
        sheet.cell(row=amort_start + year, column=3, value=interest)
        sheet.cell(row=amort_start + year, column=4, value=principal)
        sheet.cell(row=amort_start + year, column=5, value=balance)

    # Auto-adjust column widths
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20

    # Save Workbook
    workbook.save("BRRRR_Property_Calculator.xlsx")
    print("BRRRR Property Calculator has been created!")

if __name__ == "__main__":
    # Input data
    data = {
        'Purchase Price': 1050000,
        'Down Payment %': 20,
        'Interest Rate': 3.4,
        'Carrying Cost Months': 8,
        'Closing Cost': 21000,
        'Renovation Cost': 128000,
        'Post Reno Value': 1350000,
        'Refinance LTV': 80,
        'Amortization Years': 5,
        'Appreciation Rate': 8,
        'Rent Income': {"Rent 1": 3200, "Parking": 350, "Laundry": 0},
        'Operating Expenses': {
            "Hydro": 50, "Water": 50, "Gas": 100, "Insurance": 150, "Property Tax": 291,
            "Vacancy": 388, "Maintenance": 388, "Property Management": 775
        }
    }
    create_brrrr_excel(data)
